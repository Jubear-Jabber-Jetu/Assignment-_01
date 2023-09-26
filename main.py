import openpyxl
import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time

chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument("executable_path=C:\\Users\\Walton\\OneDrive\\Desktop\\Test_Folder\\chrome-win64\\chromedriver.exe")
driver = webdriver.Chrome(options=chrome_options)

# Open Google Chrome
driver.get("https://www.google.com")
# Specify the path to your Excel file
excel_file_path = r'C:\Users\Walton\OneDrive\Desktop\Excel.xlsx'  # Replace with your Excel file path

try:
    # Get the current day name
    current_day_name = datetime.datetime.now().strftime("%A")

    # Load the Excel workbook
    workbook = openpyxl.load_workbook(excel_file_path)

    # Specify the worksheet name or use the active sheet
    worksheet_name = current_day_name  # Use the current day name as the worksheet name
    try:
        worksheet = workbook[worksheet_name]
    except KeyError:
        print(f"Worksheet '{worksheet_name}' does not exist. Using the active sheet instead.")
        worksheet = workbook.active  # Use the active sheet as a fallback

    # Initialize dictionaries to store the longest and shortest suggestions for each search term
    longest_suggestions = {}
    shortest_suggestions = {}

    # Iterate through rows and get non-empty values from the 3rd column (column 'C')
    search_terms = []  # Create a list to store search terms
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=3, max_col=3):
        for cell in row:
            if cell.value is not None:
                search_terms.append(cell.value)  # Append non-empty values to the list

    # Now, iterate through the search terms and perform the searches
    for search_term in search_terms:
        wait = WebDriverWait(driver, 10)
        search_bar = wait.until(EC.visibility_of_element_located((By.NAME, "q")))

        # Enter the search term from the Excel file into the search bar
        search_bar.clear()  # Clear any existing text in the search bar
        search_bar.send_keys(search_term)

        # Wait for suggestions to appear (adjust the sleep duration as needed)
        time.sleep(2)

        # Get the suggestions from the search bar
        suggestions = driver.find_elements(By.XPATH, "//ul[@role='listbox']/li[@role='presentation']")

        # Initialize variables to store the longest and shortest suggestions
        max_length = 0
        min_length = float('inf')
        max_suggestion = ""
        min_suggestion = ""

        # Find the suggestion with the maximum and minimum text length
        for suggestion in suggestions:
            text = suggestion.text
            if len(text) > max_length:
                max_length = len(text)
                max_suggestion = text
            if len(text) < min_length:
                min_length = len(text)
                min_suggestion = text

        # Store the longest and shortest suggestions in the dictionaries
        longest_suggestions[search_term] = max_suggestion
        shortest_suggestions[search_term] = min_suggestion

    # Now, iterate through the Excel worksheet again and add the longest and shortest suggestions
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=3, max_col=3):
        for cell in row:
            search_term = cell.value
            if search_term in longest_suggestions:
                cell.offset(column=1).value = longest_suggestions[search_term]
            if search_term in shortest_suggestions:
                cell.offset(column=2).value = shortest_suggestions[search_term]

    # Save the updated Excel workbook
    workbook.save(excel_file_path)

except FileNotFoundError:
    print(f"File not found: {excel_file_path}")
except Exception as e:
    print(f"An error occurred: {str(e)}")
finally:
    # Close the WebDriver when done
    driver.quit()
